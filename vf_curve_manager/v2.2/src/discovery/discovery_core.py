"""
discovery_core.py — Platform detection, fuse-path scan, register analysis, cache
=================================================================================
Canonical home for the core discovery functions, split from the original
auto_discover_vf_registers.py monolith.

All shared module-level state (SCRIPT_DIR, _platform_config_cache, etc.) lives
here so that discovery_learn.py can import from it.
"""

from __future__ import annotations  # lazy annotation evaluation — no runtime type errors

from pkgutil import ModuleInfo
import logging
import sys
import subprocess
import shutil
import copy
import time
import json
import re
from collections import defaultdict, Counter
from pathlib import Path
from datetime import datetime
import io
import os
import traceback
import warnings

# Only override stdout encoding when running as a script, not when imported
# as a module (e.g. from unit tests). Replacing sys.stdout at import time
# breaks pytest's capture mechanism because pytest substitutes sys.stdout
# with an internal object that has no .buffer attribute.
if __name__ == '__main__' and hasattr(sys.stdout, 'buffer'):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
SCRIPT_DIR: Path = Path(__file__).parent.parent  # points to src/
PLATFORM_CONFIG_PATH: Path = SCRIPT_DIR / "platform_config.json"
# All output files (Excel export, txt report, JSON template) go to the
# project-root Logs/ directory so they appear alongside VF curve exports.
_LOGS_ROOT: Path = SCRIPT_DIR.parent / 'Logs'
_LOGS_ROOT.mkdir(exist_ok=True)


class _SuppressHWNoise:
    """Suppress all pysvtools console noise during hardware access.

    Stacks two layers so both WCL and NVL are covered:
      1. contextlib.redirect_stdout/stderr  — catches Python-level sys.stdout/
         sys.stderr writes (sufficient on WCL / Python 3.10 where pysvtools
         uses the same CRT as Python).
      2. os.dup2 fd redirect to os.devnull  — catches C-extension writes that
         bypass sys.stdout entirely (required on NVL / Python 3.13 where
         pysvtools DLLs use a different MSVC CRT instance).
    """
    def __init__(self) -> None:
        self._saved: dict[int, int] = {}
        self._null_fd: int | None = None
        self._ctx: "contextlib.ExitStack | None" = None

    def __enter__(self) -> "_SuppressHWNoise":
        import contextlib as _cl
        import io as _io
        # Layer 1: Python-level stream redirect
        self._ctx = _cl.ExitStack()
        self._ctx.__enter__()
        self._ctx.enter_context(_cl.redirect_stdout(_io.StringIO()))
        self._ctx.enter_context(_cl.redirect_stderr(_io.StringIO()))
        # Layer 2: fd-level redirect (suppresses C-extension / different-CRT writes)
        try:
            self._null_fd = os.open(os.devnull, os.O_WRONLY)
            for fd in (1, 2):
                self._saved[fd] = os.dup(fd)
                os.dup2(self._null_fd, fd)
        except OSError:
            pass  # fd redirect unavailable — layer 1 still active
        return self

    def __exit__(self, *_):
        # Restore fd-level first (so streams are live again before Python teardown)
        for fd, saved in self._saved.items():
            try:
                os.dup2(saved, fd)
                os.close(saved)
            except OSError:
                pass
        if self._null_fd is not None:
            try:
                os.close(self._null_fd)
            except OSError:
                pass
        self._saved.clear()
        self._null_fd = None
        # Restore Python-level streams
        if self._ctx is not None:
            self._ctx.__exit__(None, None, None)
            self._ctx = None


DISCOVERY_CACHE_PATH: Path = SCRIPT_DIR / 'vf_discovery_cache.json'

# ---------------------------------------------------------------------------
# Fuse-root static fallback candidates — tried in order when dynamic
# namednodes enumeration yields nothing.
# ---------------------------------------------------------------------------
_FUSE_ROOT_CANDIDATES: list[str] = [
    'cdie.fuses',   # current Intel default
    'soc.fuses',    # possible future Intel / non-Intel
    'die.fuses',
    'chip.fuses',
    'fuses',        # bare root fallback
]

# ---------------------------------------------------------------------------
# VF register keyword filter & category map.
#
# Used in two places:
#   1. get_vf_registers_in_path()  — keeps only attrs that match ANY keyword
#   2. categorize_register()       — first matching category wins
#
# Key   = category string stored in the cache 'category' field
# Value = list of lowercase substrings; any match in (name + description)
#         qualifies a register for that category.
#
# Order matters for categorize_register: more-specific entries must come
# before broader ones (e.g. 'itd_voltage' before 'voltage').
# ---------------------------------------------------------------------------
VF_KEYWORDS: dict[str, list[str]] = {
    # ── Scalar modifiers — matched first to avoid overlap with vf_curve ──
    # ITD voltage cutoff / floor (e.g. itd_cutoff_v_0, gt_itd_floor_v)
    'itd_voltage':  ['itd_cutoff_v', 'itd_floor_v', 'itd_v_'],
    'itd_slope':    ['itd_slope'],
    # P0 overrides: AVX/TMUL per-core delta ratios (NOT the VF table ratio)
    'p0_override':  ['p0_ratio_avx', 'p0_ratio_tmul', 'ia_p0_ratio_avx',
                     'ia_p0_ratio_tmul'],
    'acode_min':    ['acode_ia_min', 'acode_min'],
    'downbin':      ['_downbin'],
    'atom_delta':   ['_atom_delta'],
    'mct_delta':    ['_bigcore_delta'],

    # ── VF working-point table registers ─────────────────────────────────
    # All carry _vf_ in their names:  fw_fuses_X_vf_voltage_N,
    #   fw_fuses_X_vf_ratio_N, fw_fuses_X_vf_voltage_reg_adder_N,
    #   fw_fuses_X_vf_voltage_delta_idxN_N, and core_fuse acode/ia_base.
    'vf_curve':     ['_vf_voltage', '_vf_ratio', 'vf_voltage_', 'vf_ratio_',
                     'core_fuse_acode_ia'],

    # ── Standalone frequency ratio registers ─────────────────────────────
    # e.g. fw_fuses_ia_p0_ratio, fw_fuses_ring_p0_ratio (no _vf_ in name)
    'frequency':    ['_p0_ratio', '_p1_ratio', '_pn_ratio', '_min_ratio',
                     '_max_ratio', 'ia_p0_ratio', 'ia_min_ratio',
                     'ia_pn_ratio', 'ia_p1_ratio'],

    # ── Curve configuration (num working points etc.) ─────────────────
    'curve_config': ['num_of_points', 'num_of_vf', 'wp_count'],

    # ── Standalone voltage / power / thermal (not VF table) ──────────
    'voltage':      ['vcc', 'vnn', 'qclk_volt', 'svid_volt'],
    'power':        ['power_limit', 'tdp', '_pl1_', '_pl2_', 'pkg_power'],
    'thermal':      ['thermal', 'temp_', 'tcc', 'prochot'],
    'fivr':         ['fivr', 'dlvr', 'vrci'],

    # ── Broad catch-all for any remaining fw_fuses_ registers ────────
    'fw_fuses':     ['fw_fuses_'],
}

# ---------------------------------------------------------------------------
# Registers for which value == 0 is a legitimate (not "unprogrammed") state.
# Used by _is_zero_valid() → get_register_info() → active flag.
# ---------------------------------------------------------------------------
_ZERO_VALID_PATTERNS: tuple[str, ...] = (
    'adder', 'delta', 'vf_index', 'num_of', 'v_gap', 'vfloor', 'vceil',
)

# ---------------------------------------------------------------------------
# platform_config.json cache — avoids repeated disk reads during startup.
# _persist_learned_patterns() must call _invalidate_platform_config_cache()
# after every write so subsequent reads pick up the saved data.
# ---------------------------------------------------------------------------
_platform_config_cache: dict | None = None

# ---------------------------------------------------------------------------
# Cold-reset detective — discovery phase
# ---------------------------------------------------------------------------
# Updated by analyze_fuse_path() before every register read so that when a
# 0x8000000f "target powered down" error fires we can instantly print which
# path + register was being accessed at the moment of the cold reset.
_LAST_DISCOVERY_ACCESS: dict = {
    'fuse_path': None,
    'register':  None,
    'timestamp': None,
}

# Single source of truth — imported from utils.constants (no circular import).
from utils.constants import TARGET_DOWN_KEYWORDS as _DISCOVERY_TARGET_DOWN_KW

# Module logger — used by all functions in this file and imported by discovery_learn
log = logging.getLogger(__name__)


def _read_platform_config_json() -> dict:
    """Read and cache platform_config.json.  Returns {} on missing/error."""
    global _platform_config_cache
    if _platform_config_cache is not None:
        return _platform_config_cache
    try:
        if PLATFORM_CONFIG_PATH.exists():
            with open(PLATFORM_CONFIG_PATH, 'r', encoding='utf-8') as f:
                _platform_config_cache = json.load(f)
            return _platform_config_cache
    except Exception:
        pass
    _platform_config_cache = {}
    return _platform_config_cache


def _invalidate_platform_config_cache() -> None:
    """Invalidate the in-memory cache after any write to platform_config.json."""
    global _platform_config_cache
    _platform_config_cache = None


def _get_pythonsv_settings() -> dict:
    """Read pythonsv paths from platform_config.json, with sensible defaults."""
    defaults: dict[str, str] = {
        'config_file': 'C:/pythonsv/pysv_config.ini',
        'project_root': 'C:/pythonsv',
    }
    try:
        data = _read_platform_config_json()
        return {**defaults, **data.get('pythonsv', {})}
    except Exception:
        pass
    return defaults


def _get_platform_keys() -> list:
    """Return all platform keys defined in platform_config.json platforms dict."""
    try:
        data = _read_platform_config_json()
        keys = list(data.get('platforms', {}).keys())
        return [k for k in keys if k != 'generic']
    except Exception:
        pass
    return ['wildcatlake', 'meteorlake', 'lunarlake', 'arrowlake', 'pantherlake',
            'raptorlake', 'alderlake']


def detect_platform_name() -> str:
    """Auto-detect platform using 3-tier strategy, most to least authoritative:

      Tier 1 — pysv_config.ini  [baseaccess] project=
               Written by PythonSV itself when the project is set up.
               Path is configurable in platform_config.json pythonsv.config_file.
               e.g.  [baseaccess]\n               project = wildcatlake  ->  'wildcatlake'

      Tier 2 — pythonsv project root subdirectories
               Scans platform_config.json pythonsv.project_root for folder names
               that match a known platform key.
               e.g.  C:/pythonsv/wildcatlake/  ->  'wildcatlake'

      Tier 3 — installed Python package names
               Matches package names against platform keys (exact, prefix, substring).
               e.g.  package 'wildcatlake'  ->  'wildcatlake'
    """
    psv = _get_pythonsv_settings()
    platform_keys = _get_platform_keys()

    # --- Tier 1: pysv_config.ini [baseaccess] project= ---
    try:
        config_path = Path(psv['config_file'])
        if config_path.exists():
            import configparser
            cfg = configparser.ConfigParser()
            cfg.read(config_path)
            project: str = cfg.get('baseaccess', 'project', fallback='').strip().lower()
            if project and project != 'auto':
                log.info(f"    Source: pysv_config.ini  ({config_path})")
                return project
    except Exception:
        pass

    # --- Tier 2: pythonsv project root subdirectory names ---
    try:
        root = Path(psv['project_root'])
        if root.exists():
            subdirs: set[str] = {d.name.lower() for d in root.iterdir() if d.is_dir()}
            for key in platform_keys:
                if key in subdirs:
                    log.info(f"    Source: pythonsv root dir  ({root / key})")
                    return key
    except Exception:
        pass

    # --- Tier 3: installed package names ---
    try:
        import pkgutil
        installed: list[str] = [m.name.lower() for m in pkgutil.iter_modules()]
        # exact
        for name in installed:
            if name in platform_keys:
                log.info(f"    Source: installed package  ({name})")
                return name
        # prefix
        for name in installed:
            for key in platform_keys:
                if name.startswith(key + '_') or name.startswith(key + '-'):
                    log.info(f"    Source: installed package prefix  ({name})")
                    return key
        # substring
        for name in installed:
            for key in platform_keys:
                if key in name:
                    log.info(f"    Source: installed package substring  ({name})")
                    return key
    except Exception:
        pass

    return 'generic'


def load_platform_config(platform_name: str) -> dict:
    """Load platform config from platform_config.json.
    Falls back to generic if the platform is not listed, or to a built-in
    minimal config if the file is missing entirely.
    """
    try:
        all_configs = _read_platform_config_json()
        if all_configs:
            platforms = all_configs.get('platforms', {})
            if platform_name in platforms:
                cfg = platforms[platform_name]
                log.info(f"Loaded platform config: {cfg.get('display_name', platform_name)}")
                return cfg
            elif 'generic' in platforms:
                log.error(f"'{platform_name}' not in platform_config.json — using generic")
                return platforms['generic']
    except Exception as e:
        log.error(f"Warning: Could not load {PLATFORM_CONFIG_PATH.name}: {e}")

    # Built-in fallback — works even without platform_config.json
    log.error(f"Using built-in generic config (platform_config.json not found)")
    return {
        "display_name": f"Generic ({platform_name})",
        "fuse_root": "cdie.fuses",
        "core_fuse_pattern": "core{n}_fuse",
        "system_fuse_names": ["punit_fuses"],
        "extra_fuse_names": [],
        "bigcore_fuse_override": "core_fuse",
        "domain_patterns": {
            "bigcore": ["_ia_core", "_ia_vf", "_ia_base_vf", "_ia_p0_", "_ia_p1_", "_ia_pn_",
                        "_ia_min_ratio", "_ia_max_ratio", "_ia_reference", "_ia_itd",
                        "_ia_leak", "_ia_num_of", "_ia_delta_vf"],
            "atom":    ["atom_fuse", "ecore", "cpu0_fuse", "cpu1_fuse", "atom_ia",
                        "atom_vf", "lmt_vf", "atom_num_of"],
            "ring":    ["ring_vf", "_ring_", "ring_fuse", "llc", "cbo"],
            "gt":      ["_gt_vf", "_gt_fuse", "_gt_", "graphics", "gpu"],
            "media":   ["media_vf", "media_fuse", "vebox", "_vd_fuse"],
            "sa":      ["system_agent", "uncore_sa", "sa_fuse", "sa_vf"],
            "io":      ["pcie", "io_fuse", "_io_vf", "io_vf"],
        },
        "desc_hints": {
            "bigcore": ["performance core", "p-core", "big core"],
            "atom":    ["efficiency core", "e-core", "atom core"],
            "ring":    ["ring bus", "last level cache", "ring domain"],
            "gt":      ["graphics domain", "gpu domain", "graphics tile"],
            "sa":      ["system agent", "uncore voltage"],
            "io":      ["pcie domain", "io domain"],
        },
    }


def resolve_object(path_str: str):
    """Resolve a dot-notation path like 'cdie.fuses.core0_fuse' to an object.

    Search order for the root name:
      1. namednodes attributes  (standard pythonsv path)
      2. This module's globals  (standalone __main__ run)
      3. __main__ module globals  (embedded call — vf_curve_manager.py has
         `cdie` etc. injected by `from pysvtools.pmext.services.regs import *`)
      4. Call-stack frame globals  (any intermediate launcher module)
      5. eval() as last resort
    """
    import sys as _sys

    parts: list[str]    = path_str.split('.')
    root: str     = parts[0]
    root_obj = None

    # 1. namednodes
    try:
        import namednodes as _nn
        if hasattr(_nn, root):
            root_obj = getattr(_nn, root)
    except Exception:
        pass

    # 2. this module's own globals
    if root_obj is None:
        root_obj = globals().get(root)

    # 3. __main__ globals  ← key for embedded-call scenario
    if root_obj is None:
        try:
            _main: sys.ModuleType | None = _sys.modules.get('__main__')
            if _main is not None and hasattr(_main, root):
                root_obj = getattr(_main, root)
        except Exception:
            pass

    # 4. walk call-stack frames
    if root_obj is None:
        try:
            _frame: sys.FrameType = _sys._getframe(1)
            while _frame is not None:
                if root in _frame.f_globals:
                    root_obj = _frame.f_globals[root]
                    break
                _frame: sys.FrameType | None = _frame.f_back
        except Exception:
            pass

    # 5. eval fallback
    if root_obj is None:
        try:
            root_obj = eval(root)
        except Exception:
            pass

    if root_obj is None:
        return None

    # Navigate remaining parts via getattr chain
    try:
        obj = root_obj
        for part in parts[1:]:
            obj = getattr(obj, part)
        return obj
    except Exception:
        return None


def _enumerate_fuse_roots() -> list:
    """Return ALL live fuse roots visible on the currently connected platform.

    Discovery strategy (most to least authoritative):

      Step 1 — namednodes inspection
        namednodes is the pythonsv global namespace that holds every named node
        (cdie, cdie0, cdie1, soc, …).  We dir() it, skip private names, and
        check whether each node has a 'fuses' attribute.  This handles any
        naming convention without code changes.

      Step 2 — static candidate list
        If namednodes is unavailable or yields nothing, fall through to
        _FUSE_ROOT_CANDIDATES and return whichever paths resolve.

    Returns a list of dot-paths such as ['cdie.fuses', 'cdie0.fuses'].
    Never returns an empty list — falls back to ['cdie.fuses'] as the
    ultimate default so callers always have something to try.
    """
    found = []

    # --- Step 1: namednodes ---
    try:
        import namednodes as _nn
        for node_name in sorted(dir(_nn)):
            if node_name.startswith('_'):
                continue
            try:
                node = getattr(_nn, node_name, None)
            except Exception:
                continue
            if node is None or callable(node):
                continue
            if hasattr(node, 'fuses'):
                fuse_root: str = f"{node_name}.fuses"
                obj = resolve_object(fuse_root)
                if obj is not None:
                    found.append(fuse_root)
    except Exception:
        pass  # namednodes not available — fall through

    if found:
        return found

    # --- Step 2: static candidates ---
    for candidate in _FUSE_ROOT_CANDIDATES:
        obj = resolve_object(candidate)
        if obj is not None:
            found.append(candidate)

    return found if found else [_FUSE_ROOT_CANDIDATES[0]]


def _probe_fuse_root() -> str:
    """Return the first live fuse root (backward-compatible single-root helper)."""
    roots = _enumerate_fuse_roots()
    if roots:
        log.info(f"    [*] Auto-probed fuse root: {roots[0]}")
    return roots[0]


# ---------------------------------------------------------------------------
# Public namednodes probe — used by the probe-platform CLI command
# ---------------------------------------------------------------------------

def probe_namednodes() -> dict:
    """Enumerate every named node visible in the live namednodes namespace.

    Returns a dict:
    {
      'namednodes_available': bool,      # False if the package is not installed
      'all_nodes': [str, ...],           # every public top-level node name
      'fuse_roots': [str, ...],          # nodes that have a .fuses attribute,
                                         #   as dot-paths e.g. 'cdie.fuses'
      'fuse_containers': {               # fuse containers one level below .fuses
        'cdie.fuses': ['punit_fuses', 'core0_fuse', ...],
        ...
      },
      'node_attrs': {                    # shallow public attribute list per node
        'cdie': ['fuses', 'tap', ...],
        ...
      },
      'error': str | None,              # set when namednodes raises an exception
    }
    """
    result: dict = {
        'namednodes_available': False,
        'all_nodes': [],
        'fuse_roots': [],
        'fuse_containers': {},
        'node_attrs': {},
        'error': None,
    }

    try:
        import namednodes as _nn
        result['namednodes_available'] = True
    except ImportError as exc:
        result['error'] = f"namednodes not installed: {exc}"
        return result
    except Exception as exc:
        result['error'] = f"namednodes import error: {exc}"
        return result

    try:
        all_nodes = [n for n in sorted(dir(_nn)) if not n.startswith('_')]
        result['all_nodes'] = all_nodes
    except Exception as exc:
        result['error'] = f"dir(namednodes) failed: {exc}"
        return result

    for node_name in all_nodes:
        try:
            node = getattr(_nn, node_name, None)
        except Exception:
            continue
        if node is None or callable(node):
            continue

        # Collect shallow public attributes for this node
        try:
            attrs = [a for a in sorted(dir(node))
                     if not a.startswith('_') and not callable(getattr(node, a, None))]
            result['node_attrs'][node_name] = attrs
        except Exception:
            result['node_attrs'][node_name] = []

        # Check for .fuses
        if hasattr(node, 'fuses'):
            fuse_root = f"{node_name}.fuses"
            fuse_obj = resolve_object(fuse_root)        # use resolve_object for safety
            if fuse_obj is not None:
                result['fuse_roots'].append(fuse_root)
                try:
                    containers = [
                        a for a in sorted(dir(fuse_obj))
                        if not a.startswith('_')
                        and not callable(getattr(fuse_obj, a, None))
                    ]
                    result['fuse_containers'][fuse_root] = containers
                except Exception:
                    result['fuse_containers'][fuse_root] = []

    return result


def _discover_fuse_paths_from_config(cfg: dict, fuse_root: str, root_obj) -> list:
    """Legacy config-driven fuse path discovery — used as fallback only.

    Adds system_fuse_names, probes core{n}_fuse until absent,
    then adds extra_fuse_names.  Returns an ordered list of dot-paths.
    """
    core_pattern = cfg.get('core_fuse_pattern', 'core{n}_fuse')
    system_names = cfg.get('system_fuse_names', [])
    extra_names  = cfg.get('extra_fuse_names', [])

    paths = []

    for name in system_names:
        if hasattr(root_obj, name):
            paths.append(f"{fuse_root}.{name}")

    core_num = 0
    while True:
        core_name = core_pattern.format(n=core_num)
        if hasattr(root_obj, core_name):
            paths.append(f"{fuse_root}.{core_name}")
            core_num += 1
        else:
            break

    for name in extra_names:
        if hasattr(root_obj, name):
            p: str = f"{fuse_root}.{name}"
            if p not in paths:
                paths.append(p)

    return paths


def _enumerate_containers_under_root(fuse_root: str, root_obj, cfg: dict) -> list:
    """Return (attr_name, full_path, bucket) triples for every fuse container
    directly under *root_obj* (i.e. one level below the fuse root).

    Bucket values: 'system', 'core  ', 'other '
    Returned order: system → cores (numerically sorted) → other.
    Returns an empty list if nothing usable is found.
    """
    core_pattern = cfg.get('core_fuse_pattern', 'core{n}_fuse')
    system_cfg   = {n.lower() for n in cfg.get('system_fuse_names', [])}
    extra_cfg    = {n.lower() for n in cfg.get('extra_fuse_names', [])}

    core_re_src = re.escape(core_pattern).replace(r'\{n\}', r'\d+') + '$'
    core_re = re.compile(core_re_src, re.IGNORECASE)

    system_paths: list = []
    core_paths:   list = []
    other_paths:  list = []

    for attr_name in sorted(dir(root_obj)):
        if attr_name.startswith('_'):
            continue
        try:
            child = getattr(root_obj, attr_name, None)
        except Exception:
            continue
        if child is None or callable(child):
            continue
        try:
            child_regs: list[str] = [
                x for x in dir(child)
                if not x.startswith('_') and not callable(getattr(child, x, None))
            ]
        except Exception:
            child_regs = []
        if not child_regs:
            continue

        full_path: str  = f"{fuse_root}.{attr_name}"
        attr_lower: str = attr_name.lower()

        if attr_lower in system_cfg or attr_lower in extra_cfg:
            system_paths.append((attr_name, full_path, 'system'))
        elif core_re.match(attr_name):
            core_paths.append((attr_name, full_path, 'core  '))
        else:
            other_paths.append((attr_name, full_path, 'other '))

    def _core_key(item: tuple) -> int:
        m: re.Match[str] | None = re.search(r'\d+', item[0])
        return int(m.group()) if m else 0

    core_paths.sort(key=_core_key)
    return system_paths + core_paths + other_paths


def discover_fuse_paths(cfg: dict) -> list:
    """Discover ALL fuse containers across ALL live fuse roots on this platform.

    Fuse root enumeration strategy (most to least authoritative):

      1. namednodes inspection — dir() the pythonsv global namespace, find
         every top-level node (cdie, cdie0, cdie1, soc, …) that exposes a
         .fuses attribute.  Handles any naming convention automatically.
      2. Static fallback list — _FUSE_ROOT_CANDIDATES tried in order
         (cdie.fuses, cdie0.fuses, soc.fuses, …).

    For each live fuse root, dir() its .fuses object and classify every
    non-private, non-callable container child as:

      [system]  — listed in config system_fuse_names / extra_fuse_names
      [core  ]  — matches core_fuse_pattern (e.g. core0_fuse … core_fuse)
      [other ]  — everything else (gt_fuse, sa_fuse, media_fuse, …)

    Paths from all roots are aggregated and returned ordered:
      system → cores → other  (within each root, roots sorted by name).

    Falls back to the legacy config-driven probe if dir() yields nothing.
    """
    # --- Enumerate all live fuse roots ---
    configured_root = cfg['fuse_root']
    all_roots = _enumerate_fuse_roots()

    # Ensure the configured root is tried even if not found by namednodes
    if configured_root not in all_roots:
        obj = resolve_object(configured_root)
        if obj is not None:
            all_roots.insert(0, configured_root)

    if not all_roots:
        log.error(f"No live fuse roots found — discovery aborted")
        return []

    log.info(f"\n[*] Found {len(all_roots)} live fuse root(s): {all_roots}")
    # Update cfg so downstream steps (load_fuse_ram, report) use the first live root
    cfg['fuse_root'] = all_roots[0]

    all_paths:   list = []
    seen_paths: set  = set()

    for fuse_root in all_roots:
        root_obj = resolve_object(fuse_root)
        if root_obj is None:
            log.error(f"    [!] {fuse_root} — could not resolve, skipping")
            continue

        containers = _enumerate_containers_under_root(fuse_root, root_obj, cfg)

        if not containers:
            log.error(f"    [!] {fuse_root} — dir() found no containers, "
                  "trying config-driven probe")
            fallback = _discover_fuse_paths_from_config(cfg, fuse_root, root_obj)
            for p in fallback:
                if p not in seen_paths:
                    all_paths.append(p)
                    seen_paths.add(p)
            continue

        log.info(f"\n[+] {fuse_root} — {len(containers)} container(s):")
        for attr_name, full_path, bucket in containers:
            log.info(f"    [{bucket}]  {full_path}")
            if full_path not in seen_paths:
                all_paths.append(full_path)
                seen_paths.add(full_path)

    return all_paths


def load_fuse_ram_once(fuse_root: str) -> bool:
    """Load fuse RAM from the fuse root. Only needs to be done once per session.

    Skips the load when pysvtools has already loaded fuse RAM during ITP init
    (detectable via the private _fuse_ram_loaded flag or by probing a register
    value without triggering a hardware read).  This avoids a double ~5-min
    load when called from the embedded pipeline.
    """
    try:
        obj = resolve_object(fuse_root)
        if obj is None:
            log.error(f"Cannot resolve fuse root: {fuse_root}")
            return False

        # Check if fuse RAM is already loaded — pysvtools sets internal flags
        # or the object exposes a 'loaded' / '_loaded' attribute after load.
        already_loaded = False
        for flag in ('_fuse_ram_loaded', 'fuse_ram_loaded', '_loaded', 'loaded'):
            val = getattr(obj, flag, None)
            if val is True:
                already_loaded = True
                break

        def _notify_hw_access_loaded() -> None:
            """Tell hardware_access.py that this root path is loaded so it
            skips redundant re-loads (which would re-trigger the postcondition
            and risk a cold reset in active-boot state)."""
            try:
                from utils.hardware_access import notify_fuse_ram_loaded
                notify_fuse_ram_loaded(fuse_root)
            except Exception:
                pass  # non-fatal if hw module not available

        if already_loaded:
            log.info(f"Fuse RAM already loaded (skipping re-load)")
            _notify_hw_access_loaded()
            return True

        if not hasattr(obj, 'load_fuse_ram'):
            log.error("load_fuse_ram() not available on this platform")
            return False

        # Timeout guard: platforms like NovaLake (NVL) with multiple tiles can
        # cause load_fuse_ram() to hang indefinitely.  Run it in a daemon thread
        # with a generous timeout so discovery is never permanently blocked.
        _LOAD_FUSE_RAM_TIMEOUT_SEC: int = 720   # 12 minutes

        log.info(f"\n[*] Loading fuse RAM from {fuse_root} "
                 f"(up to {_LOAD_FUSE_RAM_TIMEOUT_SEC // 60} min, please wait)...")
        print(f"    [*] Loading fuse RAM from {fuse_root} "
              f"(up to {_LOAD_FUSE_RAM_TIMEOUT_SEC // 60} min)...",
              flush=True)
        start: float = time.time()

        def _do_load() -> None:
            # Use fd-level suppression (_SuppressHWNoise) so that C-extension code in
            # pysvtools (Python 3.13 on NVL) which writes directly to fd-1/fd-2 is also
            # silenced.  contextlib.redirect_stdout only hooks sys.stdout and would miss
            # the 'post condition failed' / AccessTimeoutError traceback on NVL.
            with _SuppressHWNoise():
                obj.load_fuse_ram()

        import concurrent.futures as _cf
        try:
            with _cf.ThreadPoolExecutor(max_workers=1) as _pool:
                _fut = _pool.submit(_do_load)
                _fut.result(timeout=_LOAD_FUSE_RAM_TIMEOUT_SEC)
        except _cf.TimeoutError:
            elapsed = time.time() - start
            log.warning(
                f"\n[!] load_fuse_ram() timed out after {elapsed/60:.1f} min on '{fuse_root}'.\n"
                f"    This is common on multi-tile platforms (NVL, PTL) where ITP may\n"
                f"    pre-load fuse RAM during tap-unlock.  Continuing as-if loaded."
            )
            print(
                f"    [!] load_fuse_ram() timed out after {elapsed/60:.1f} min — "
                f"assuming pre-loaded by ITP.", flush=True
            )
            _notify_hw_access_loaded()
            return True

        log.info(f"    Fuse RAM loaded in {time.time() - start:.1f} seconds\n")
        print(f"    [+] Fuse RAM loaded in {time.time() - start:.1f} s", flush=True)
        _notify_hw_access_loaded()
        return True
    except Exception as e:
        error_str: str = str(e).lower()
        exc_type: str  = type(e).__name__.lower()
        # Postcondition failure: "post condition failed" / AccessTimeoutError in
        # the cleanup phase.  The fuse data IS in memory — only the _enable_dcg
        # IOSF-SB write failed.  Treat as loaded so the UI doesn't re-load and
        # re-trigger the same write in active-boot state (which cold-resets).
        # NOTE: str(e) contains the timeout *message*, not the class name, so
        # we must also check type(e).__name__ to catch AccessTimeoutError.
        is_postcondition: bool = ('post condition' in error_str or
                            'postcondition' in error_str or
                            'accesstimeouterror' in error_str.replace(' ', '') or
                            'accesstimeouterror' in exc_type)
        if is_postcondition:
            log.warning(
                f"\n[!] Fuse RAM post-condition timed out for '{fuse_root}' "
                f"— this is expected on active-boot platforms.\n"
                f"    The fuse data IS fully loaded in memory; continuing normally.")
            try:
                from utils.hardware_access import notify_fuse_ram_loaded
                notify_fuse_ram_loaded(fuse_root)
            except Exception:
                pass
            return True
        if any(kw in error_str for kw in _DISCOVERY_TARGET_DOWN_KW):
            log.info("")
            log.info("!" * 70)
            log.info("  [COLD-RESET-DETECTIVE] Target reset during load_fuse_ram_once()")
            log.info(f"  fuse_root : {fuse_root}")
            log.info(f"  Error     : {e}")
            log.info("  TIP: the fuse RAM load itself is triggering the cold reset.")
            log.info("!" * 70)
            log.info("")
        else:
            log.error(f"ERROR loading fuse RAM: {e}")
        return False


def get_vf_registers_in_path(path_str: str) -> tuple:
    """Return (obj, [all_register_names]) for a fuse path.

    Returns every non-private, non-callable attribute on the fuse container —
    not just those matching VF_KEYWORDS.  Categorization (vf_curve, frequency,
    voltage, other, …) is deferred to categorize_register() so that registers
    whose names do not match any pre-defined keyword are still captured with
    category='other' rather than silently dropped.
    """
    obj = resolve_object(path_str)
    if obj is None:
        return None, []

    all_attrs: list[str] = dir(obj)
    register_names: list[str] = [
        a for a in all_attrs
        if not a.startswith('_') and not callable(getattr(obj, a, None))
    ]
    return obj, register_names


def _is_zero_valid(reg_name: str) -> bool:
    """Return True when 0 is a legitimate programmed value for this register.

    For most VF fuse registers a value of 0 means the fuse was never blown
    (i.e. unprogrammed).  However adder / delta / index registers legitimately
    hold 0 to mean "no adjustment" or "first entry", so they must not be
    treated as inactive just because their value is 0.
    """
    name: str = reg_name.lower()
    return any(pat in name for pat in _ZERO_VALID_PATTERNS)


def get_register_info(obj, reg_name: str) -> dict:
    """Read value and description for one register."""
    info = {
        'name': reg_name,
        'value': None,
        'hex': None,
        'description': None,
        'accessible': False,
        'active': False,
    }
    try:
        if hasattr(obj, reg_name):
            # Suppress 'virtual fuse overriding this direct fuse' warnings emitted
            # by pysvtools for every register read on NVL (200+ lines per scan).
            with warnings.catch_warnings():
                warnings.filterwarnings('ignore')
                value = getattr(obj, reg_name)
            if value is not None:
                info['value'] = int(value)
                info['hex'] = f"0x{int(value):x}"
                info['accessible'] = True
                info['active'] = (int(value) != 0) or _is_zero_valid(reg_name)

            # Description — try multiple sources
            try:
                attr = getattr(type(obj), reg_name, None)
                if attr and hasattr(attr, '__doc__') and attr.__doc__:
                    info['description'] = attr.__doc__.strip()
            except Exception:
                pass
            try:
                reg_obj = getattr(obj, reg_name)
                if hasattr(reg_obj, 'comment') and reg_obj.comment:
                    info['description'] = reg_obj.comment
                elif hasattr(reg_obj, 'description') and reg_obj.description:
                    info['description'] = reg_obj.description
            except Exception:
                pass
    except Exception as e:
        info['error'] = str(e)
        # ── Cold-reset detective ──────────────────────────────────────────
        error_str: str = str(e).lower()
        if any(kw in error_str for kw in _DISCOVERY_TARGET_DOWN_KW):
            lda = _LAST_DISCOVERY_ACCESS
            log.info("")
            log.info("!" * 70)
            log.info("  [COLD-RESET-DETECTIVE] Target reset during register read!")
            log.info(f"  fuse_path : {lda.get('fuse_path', '?')}")
            log.info(f"  register  : {lda.get('register',  '?')}  <-- LIKELY CULPRIT")
            log.info(f"  timestamp : {lda.get('timestamp', '?')}")
            log.info(f"  Error     : {e}")
            log.info("  TIP: note the register name above and cross-reference with")
            log.info("       the domain listed in vf_domains.json.")
            log.info("!" * 70)
            log.info("")
    return info


def categorize_register(reg_name: str, description: str, cfg: dict) -> tuple:
    """Return (category, domain) using platform config patterns.

    - Category: matched against name + description (categories are universal)
    - Domain:   matched against register NAME only first (avoids false positives
                like 'io' in 'ratio', 'sa' in 'versa', 'efficiency' in FIVR desc)
                Falls back to description with config-defined multi-word hints.
    """
    reg_lower: str = reg_name.lower()
    desc_lower: str = (description or '').lower()
    combined: str = reg_lower + ' ' + desc_lower

    # --- Category ---
    category = 'other'
    for cat, keywords in VF_KEYWORDS.items():
        if any(kw in combined for kw in keywords):
            category: str = cat
            break

    # --- Domain: name-only patterns from config ---
    domain_patterns = cfg.get('domain_patterns', {})
    domain = 'unknown'
    for dom, patterns in domain_patterns.items():
        if any(pat in reg_lower for pat in patterns):
            domain = dom
            break

    # --- Domain fallback: description with multi-word hints from config ---
    if domain == 'unknown':
        for dom, phrases in cfg.get('desc_hints', {}).items():
            if any(ph in desc_lower for ph in phrases):
                domain = dom
                break

    # --- Override: core_fuse / ia_ registers belong to bigcore
    #     (corrects leftover false positives e.g. fivr, cdyn, itd) ---
    bigcore_override = cfg.get('bigcore_fuse_override', 'core_fuse')
    atom_patterns = domain_patterns.get('atom', [])
    if bigcore_override and (bigcore_override in reg_lower or 'ia_' in reg_lower):
        is_atom: bool = any(p in reg_lower for p in atom_patterns)
        if not is_atom and domain not in ('bigcore', 'atom', 'ring', 'gt', 'media'):
            domain = 'bigcore'

    return category, domain


def analyze_fuse_path(path_str: str, path_label: str, cfg: dict) -> dict:
    """Discover and categorize VF registers in one fuse path."""
    log.info(f"\n{'=' * 80}")
    log.info(f"ANALYZING: {path_label}  [{path_str}]")
    log.info(f"{'=' * 80}\n")

    obj, register_names = get_vf_registers_in_path(path_str)
    if obj is None:
        log.info(f"[x] Cannot access {path_str}")
        return {}

    log.info(f"Found {len(register_names)} registers (all categories)")
    log.info("Reading register values and descriptions...")

    results_by_category = defaultdict(list)
    global _LAST_DISCOVERY_ACCESS
    # Wrap the entire per-register scan in a single _SuppressHWNoise context so
    # that pysvtools noise (Decoding fuse ram string, Couldn't check security
    # level, post condition failed tracebacks, Couldn't find in device list)
    # is silenced for the whole path without creating a new context per register.
    with _SuppressHWNoise():
        for reg_name in register_names:
            # ── Update tracker before every hardware read ─────────────────────
            _LAST_DISCOVERY_ACCESS = {
                'fuse_path': path_str,
                'register':  reg_name,
                'timestamp': datetime.now().strftime('%H:%M:%S.%f'),
            }
            info = get_register_info(obj, reg_name)
            if info['accessible']:
                category, domain = categorize_register(reg_name, info['description'] or '', cfg)
                info['category'] = category
                info['domain'] = domain
                results_by_category[category].append(info)

    total_active: int = sum(len([r for r in regs if r['active']]) for regs in results_by_category.values())
    log.info(f"{total_active} active (non-zero) registers found in {path_label}")
    return dict(results_by_category)


def generate_recommendations(all_path_results: dict) -> dict:
    """Aggregate results from all fuse paths into prioritized recommendations.

    Note:
        **Standalone-only** — used by ``__main__`` to generate the human-
        readable text report.  This function is NOT called by the GUI or CLI
        (which use the discovery cache / vf_domains.json instead).
    """
    known_domains = _load_known_domains()
    recommendations = {d: {'high_priority': [], 'medium_priority': [], 'low_priority': []}
                       for d in known_domains + ('other_domains',)}

    for _path, results_by_cat in all_path_results.items():
        for category, registers in results_by_cat.items():
            for reg_info in registers:
                if not reg_info['active']:
                    continue

                domain = reg_info['domain']
                if domain not in recommendations:
                    domain = 'other_domains'

                if category in ('vf_curve', 'frequency', 'voltage', 'power', 'curve_config'):
                    priority = 'high_priority'
                elif category in ('thermal', 'fivr'):
                    priority = 'medium_priority'
                else:
                    priority = 'low_priority'

                recommendations[domain][priority].append({
                    'name': reg_info['name'],
                    'value': reg_info['value'],
                    'hex': reg_info['hex'],
                    'category': category,
                    'description': reg_info.get('description') or 'No description available',
                })

    return recommendations


def save_detailed_report(all_path_results: dict, recommendations: dict,
                         platform_name: str, fuse_paths: list) -> str:
    """Save human-readable analysis report.

    Note:
        **Standalone-only** — called only from ``__main__``.  Not used by the
        GUI or CLI tool at runtime.
    """
    report_path = str(_LOGS_ROOT / f"vf_register_auto_analysis_{time.strftime('%Y%m%d_%H%M%S')}.txt")

    with open(report_path, 'w', encoding='utf-8') as f:
        f.write("=" * 80 + "\n")
        f.write("AUTOMATIC VF REGISTER DISCOVERY & ANALYSIS REPORT\n")
        f.write("=" * 80 + "\n")
        f.write(f"Generated : {time.strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"Platform  : {platform_name}\n")
        f.write(f"Fuse paths: {len(fuse_paths)}\n\n")
        f.write("PURPOSE:\n")
        f.write("Automatically identifies VF-related registers, categorizes them\n")
        f.write("by domain and function, shows current values, and recommends which\n")
        f.write("registers to add to vf_domains.json.\n\n")

        # Per-path results
        for path_str, results_by_cat in all_path_results.items():
            f.write("\n" + "=" * 80 + "\n")
            f.write(f"PATH: {path_str}\n")
            f.write("=" * 80 + "\n\n")

            for category, registers in sorted(results_by_cat.items()):
                if not registers:
                    continue
                active_regs = [r for r in registers if r['active']]
                f.write(f"\n{category.upper().replace('_', ' ')}:\n")
                f.write("-" * 80 + "\n")
                f.write(f"Total: {len(registers)} | Active (non-zero): {len(active_regs)}\n\n")

                for reg in active_regs[:20]:
                    f.write(f"  {reg['name']}\n")
                    f.write(f"    Value: {reg['value']} ({reg['hex']})\n")
                    f.write(f"    Domain: {reg['domain']}\n")
                    if reg.get('description'):
                        f.write(f"    Description: {reg['description'][:200]}\n")
                    f.write("\n")

                if len(active_regs) > 20:
                    f.write(f"  ... and {len(active_regs) - 20} more active registers\n\n")

        # Recommendations
        f.write("\n" + "=" * 80 + "\n")
        f.write("INTEGRATION RECOMMENDATIONS FOR VF_DOMAINS.JSON\n")
        f.write("=" * 80 + "\n\n")

        for domain, priorities in sorted(recommendations.items()):
            if not any(priorities.values()):
                continue
            f.write(f"\n{domain.upper()}:\n{'=' * 80}\n")
            for level in ('high_priority', 'medium_priority', 'low_priority'):
                regs = priorities[level]
                if not regs:
                    continue
                f.write(f"\n{level.replace('_', ' ').upper()}:\n{'-' * 80}\n")
                f.write(f"Count: {len(regs)}\n\n")
                for reg in regs:
                    f.write(f"  Register: {reg['name']}\n")
                    f.write(f"  Value   : {reg['value']} ({reg['hex']})\n")
                    f.write(f"  Category: {reg['category']}\n")
                    f.write(f"  Desc    : {reg['description'][:150]}\n")
                    f.write(f"  Action  : Add to {domain} domain configuration\n\n")

        # Summary
        f.write("\n" + "=" * 80 + "\n")
        f.write("SUMMARY STATISTICS\n")
        f.write("=" * 80 + "\n\n")

        for path_str, results_by_cat in all_path_results.items():
            total: int = sum(len(r) for r in results_by_cat.values())
            active: int = sum(len([x for x in r if x['active']]) for r in results_by_cat.values())
            f.write(f"{path_str}:\n  Total VF-related: {total} | Active: {active}\n\n")

        total_rec: int = sum(len(r) for d in recommendations.values() for r in d.values())
        f.write(f"Total Recommendations : {total_rec}\n")
        f.write(f"  High Priority   : {sum(len(d['high_priority']) for d in recommendations.values())}\n")
        f.write(f"  Medium Priority : {sum(len(d['medium_priority']) for d in recommendations.values())}\n")
        f.write(f"  Low Priority    : {sum(len(d['low_priority']) for d in recommendations.values())}\n")

    log.info(f"Report saved to: {report_path}")
    return report_path


def save_json_template(recommendations: dict) -> str:
    """Save JSON template for easy integration into vf_domains.json.

    Note:
        **Standalone-only** — called only from ``__main__``.  Not used by the
        GUI or CLI tool at runtime.
    """
    template_path = str(_LOGS_ROOT / f"vf_domains_additions_template_{time.strftime('%Y%m%d_%H%M%S')}.json")
    template = {
        "_comment": "Recommended additions to vf_domains.json based on automatic discovery",
        "_generated": time.strftime('%Y-%m-%d %H:%M:%S'),
        "recommended_registers": {},
    }

    for domain, priorities in recommendations.items():
        if not any(priorities.values()):
            continue
        template["recommended_registers"][domain] = {
            "high_priority": [
                {"register_name": r['name'], "current_value": r['value'],
                 "category": r['category'], "description": r['description'][:100]}
                for r in priorities['high_priority']
            ],
            "medium_priority": [
                {"register_name": r['name'], "current_value": r['value'],
                 "category": r['category'], "description": r['description'][:100]}
                for r in priorities['medium_priority']
            ],
        }

    with open(template_path, 'w', encoding='utf-8') as f:
        json.dump(template, f, indent=2)

    log.info(f"JSON template saved to: {template_path}")
    return template_path


def _all_results_to_flat_records(all_path_results: dict) -> list:
    """Flatten all_path_results into a list of plain dicts for caching and export.

    Each record: name, value, hex, active, category, domain, fuse_path, description.
    Sorted: active registers first, then by domain, then alphabetically.
    """
    # Lazy import avoids circular dependency (discovery_learn imports discovery_core).
    from .discovery_learn import _infer_conversion_from_description  # noqa: PLC0415
    records = []
    for fuse_path, results_by_cat in all_path_results.items():
        for category, registers in results_by_cat.items():
            for reg in registers:
                records.append({
                    'name':        reg['name'],
                    'value':       reg.get('value'),
                    'hex':         reg.get('hex', ''),
                    'active':      (bool(reg.get('active', False)) or _is_zero_valid(reg['name'])),
                    'category':    category,
                    'domain':      reg.get('domain', 'unknown'),
                    'fuse_path':   fuse_path,
                    'description': reg.get('description') or '',
                    'converted':   _infer_conversion_from_description(
                                       reg['name'],
                                       reg.get('description') or '',
                                       reg.get('value')),
                })
    records.sort(key=lambda r: (not r['active'], r['domain'], r['name']))
    return records


def _save_discovery_cache(records: list, platform_name: str,
                          platform_display: str) -> None:
    """Write flat register snapshot to vf_discovery_cache.json."""
    try:
        # Coerce every field to a plain Python type so that ITP / pysvtools
        # integer objects (numpy.int64, custom IntField, etc.) never cause a
        # TypeError mid-dump that truncates the file and leaves it corrupt.
        clean = []
        for r in records:
            v = r.get('value')
            clean.append({
                'name':        str(r.get('name', '')),
                'value':       (int(v) if v is not None else None),
                'hex':         str(r.get('hex', '')),
                'active':      bool(r.get('active', False)),
                'category':    str(r.get('category', '')),
                'domain':      str(r.get('domain', 'unknown')),
                'fuse_path':   str(r.get('fuse_path', '')),
                'description': str(r.get('description') or ''),
                'converted':   str(r.get('converted', '')),
            })
        cache = {
            'platform':         platform_name,
            'platform_display': platform_display,
            'timestamp':        time.strftime('%Y-%m-%d %H:%M:%S'),
            'count':            len(clean),
            'registers':        clean,
        }
        with open(DISCOVERY_CACHE_PATH, 'w', encoding='utf-8') as f:
            json.dump(cache, f, indent=2, ensure_ascii=False)
        log.info(f"    [+] Discovery cache saved: {DISCOVERY_CACHE_PATH.name}"
              f"  ({len(clean)} registers)")
    except Exception as e:
        log.error(f"    [!] Could not save discovery cache: {e}")
        traceback.print_exc()


def load_discovery_cache() -> tuple:
    """Load the flat register snapshot from vf_discovery_cache.json.

    Returns:
        (records, platform_name, platform_display, timestamp)
        All values are None when the cache does not exist or is corrupt.
    """
    if not DISCOVERY_CACHE_PATH.exists():
        return None, None, None, None
    try:
        with open(DISCOVERY_CACHE_PATH, 'r', encoding='utf-8') as f:
            cache = json.load(f)
        return (
            cache.get('registers', []),
            cache.get('platform', 'generic'),
            cache.get('platform_display', 'Unknown Platform'),
            cache.get('timestamp', ''),
        )
    except Exception as e:
        log.error(f"Could not load discovery cache: {e}")
        return None, None, None, None


def save_discovery_cache_edits(edits: dict) -> int:
    """Apply metadata edits to the discovery cache JSON.

    Only the editable fields (domain, category, description, notes) are
    updated — hardware-read values (value, hex, active) are never touched.

    Args:
        edits: {register_name: {field: new_value, ...}}
               Editable fields: 'domain', 'category', 'description', 'notes'.

    Returns:
        Number of records modified, or -1 on I/O error.
    """
    _EDITABLE: set[str] = {'domain', 'category', 'description', 'notes'}
    if not DISCOVERY_CACHE_PATH.exists():
        log.error("No discovery cache to edit.")
        return 0
    try:
        with open(DISCOVERY_CACHE_PATH, 'r', encoding='utf-8') as f:
            cache = json.load(f)
        records = cache.get('registers', [])
        count = 0
        for reg in records:
            name = reg.get('name', '')
            if name in edits:
                for field, value in edits[name].items():
                    if field in _EDITABLE:
                        reg[field] = value
                count += 1
        cache['registers'] = records
        with open(DISCOVERY_CACHE_PATH, 'w', encoding='utf-8') as f:
            json.dump(cache, f, indent=2)
        return count
    except Exception as e:
        log.error(f"Could not save edits: {e}")
        return -1


def export_discovered_registers_to_excel(platform_display: str = None,
                                         records: list = None,
                                         active_only: bool = False) -> str:
    """Export discovered registers to a formatted Excel workbook.

    Filename  :  VF_Registers_{Platform}_{YYYYMMDD_HHMMSS}.xlsx
    Location  :  Logs/  (project root, same folder as all VF curve exports)
    Sheets    :
      'VF Registers'  — colour-coded register table, frozen header, auto-filter
      'Summary'       — platform info + domain / category breakdowns

    Args:
        platform_display: Human-readable platform name used in the filename.
        records:          Flat list from _all_results_to_flat_records().
                          Loaded from cache automatically when None.
        active_only:      When True, only active (non-zero) registers are written.

    Returns:
        Absolute path to the created file, or None on failure.
    """
    if records is None:
        loaded_recs, _, loaded_disp, _ = load_discovery_cache()
        if loaded_recs is None:
            log.error("No discovery cache — run discovery first.")
            return None
        records = loaded_recs
        if platform_display is None:
            platform_display = loaded_disp

    if platform_display is None:
        platform_display = 'Unknown_Platform'

    if active_only:
        records = [r for r in records if r.get('active')]

    safe_name: str = re.sub(r'[^\w\-]', '_', platform_display)
    stamp: str     = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename: str  = f"VF_Registers_{safe_name}_{stamp}.xlsx"
    filepath: Path  = _LOGS_ROOT / filename

    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter

        wb = Workbook()

        # ── Sheet 1: VF Registers ──────────────────────────────────────────
        ws = wb.active
        ws.title = 'VF Registers'

        COLS: list[str]   = ['Register Name', 'Value (Dec)', 'Value (Hex)', 'Converted (MHz / mV)',
                  'Active', 'Category', 'Domain', 'Fuse Path', 'Description']
        WIDTHS: list[int] = [52, 13, 14, 20, 8, 16, 14, 38, 80]

        hdr_fill   = PatternFill(start_color='0071C5', end_color='0071C5', fill_type='solid')
        hdr_font   = Font(color='FFFFFF', bold=True, size=11)
        act_fill   = PatternFill(start_color='E6F4EA', end_color='E6F4EA', fill_type='solid')
        inact_fill = PatternFill(start_color='FFF8F0', end_color='FFF8F0', fill_type='solid')
        wrap_aln   = Alignment(wrap_text=True, vertical='top')
        ctr_aln    = Alignment(horizontal='center', vertical='center')
        top_aln    = Alignment(vertical='center')

        for ci, (col_name, width) in enumerate(zip(COLS, WIDTHS), 1):
            cell = ws.cell(row=1, column=ci, value=col_name)
            cell.fill      = hdr_fill
            cell.font      = hdr_font
            cell.alignment = ctr_aln
            ws.column_dimensions[get_column_letter(ci)].width = width
        ws.row_dimensions[1].height = 22
        ws.freeze_panes = 'A2'

        # openpyxl only accepts int/float/str/bool/datetime/None.
        # PythonSV register values are often custom hex objects (e.g. 0x1)
        # that are not plain ints, so we must coerce them before writing.
        def _xl(v):
            """Coerce a value to an openpyxl-safe type."""
            if v is None or isinstance(v, (int, float, str, bool)):
                return v
            try:
                return int(v)           # works for PythonSV hex objects
            except (TypeError, ValueError):
                return str(v)           # last resort

        for ri, rec in enumerate(records, 2):
            is_active = bool(rec.get('active'))
            rf = act_fill if is_active else inact_fill
            raw_val = rec.get('value')
            vals = [
                rec.get('name', ''),
                _xl(raw_val),                           # Dec: int or None
                rec.get('hex', '') or (f'0x{int(raw_val):X}' if raw_val is not None else ''),
                rec.get('converted', ''),               # Converted (MHz / mV)
                'Yes' if is_active else 'No',
                rec.get('category', ''),
                rec.get('domain', ''),
                rec.get('fuse_path', ''),
                rec.get('description', ''),
            ]
            for ci, val in enumerate(vals, 1):
                cell = ws.cell(row=ri, column=ci, value=_xl(val))
                cell.fill      = rf
                cell.alignment = wrap_aln if ci == len(COLS) else top_aln

        ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"

        # ── Sheet 2: Summary ───────────────────────────────────────────────
        ws2 = wb.create_sheet(title='Summary')
        ws2['A1'] = 'VF Register Discovery Summary'
        ws2['A1'].font = Font(bold=True, size=14, color='0071C5')

        bold = Font(bold=True)
        meta = [
            ('Platform',        platform_display),
            ('Generated',       datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
            ('Total Registers', len(records)),
            ('Active (≠ 0)',    sum(1 for r in records if r.get('active'))),
            ('Inactive (= 0)',  sum(1 for r in records if not r.get('active'))),
        ]
        for ri, (k, v) in enumerate(meta, 3):
            ws2.cell(row=ri, column=1, value=k).font = bold
            ws2.cell(row=ri, column=2, value=v)

        ws2.cell(row=9, column=1, value='Domain').font   = bold
        ws2.cell(row=9, column=2, value='Count').font    = bold
        ws2.cell(row=9, column=3, value='Category').font = bold
        ws2.cell(row=9, column=4, value='Count').font    = bold

        dom_counts = Counter(r.get('domain',   'unknown') for r in records)
        cat_counts = Counter(r.get('category', 'other')   for r in records)
        for ri, (d, c) in enumerate(sorted(dom_counts.items()), 10):
            ws2.cell(row=ri, column=1, value=d)
            ws2.cell(row=ri, column=2, value=c)
        for ri, (d, c) in enumerate(sorted(cat_counts.items()), 10):
            ws2.cell(row=ri, column=3, value=d)
            ws2.cell(row=ri, column=4, value=c)

        for col_l, w in [('A', 30), ('B', 14), ('C', 20), ('D', 10)]:
            ws2.column_dimensions[col_l].width = w

        wb.save(str(filepath))
        log.info(f"Registers exported: {filepath}")
        return str(filepath)

    except Exception as e:
        log.error(f"Export failed: {e}")
        traceback.print_exc()
        return None


def export_scalar_modifiers_to_excel(platform_display: str = None,
                                      vf_domains_path: Path = None) -> str:
    """Export all discovered scalar modifier registers to a formatted Excel workbook.

    Values are cross-referenced from the discovery cache so every row shows the
    current hardware reading alongside the config metadata.

    Filename  :  Scalar_Modifiers_{Platform}_{YYYYMMDD_HHMMSS}.xlsx
    Location  :  Logs/  (project root, same folder as all VF curve exports)
    Sheets    :
      'Scalar Modifiers'  — full table: config + live value columns
      'By Type'           — grouped summary by modifier type

    Args:
        platform_display: Human-readable platform name used in the filename.
        vf_domains_path:  Path to vf_domains.json.  Defaults to canonical location.

    Returns:
        Absolute path to the created file, or None on failure.
    """
    if vf_domains_path is None:
        vf_domains_path = SCRIPT_DIR / 'vf_domains.json'

    if not vf_domains_path.exists():
        log.error("export_scalar_modifiers_to_excel: vf_domains.json not found.")
        return None

    with open(vf_domains_path, 'r', encoding='utf-8') as _f:
        vf_data = json.load(_f)

    scalars: dict = vf_data.get('scalar_modifiers', {})
    if not scalars:
        log.warning("export_scalar_modifiers_to_excel: no scalar_modifiers in vf_domains.json")
        return None

    if platform_display is None:
        platform_display = vf_data.get('platform', 'Unknown_Platform')

    # Build a fast lookup from the discovery cache so we can populate live values.
    # Cache records: {name, value, hex, converted, active, domain, category, fuse_path, description}
    cached_recs, _, cached_display, _ = load_discovery_cache()
    _cache_lookup: dict = {}
    if cached_recs:
        for r in cached_recs:
            _cache_lookup[r['name']] = r
    if platform_display is None:
        platform_display = cached_display or 'Unknown_Platform'

    safe_name: str = re.sub(r'[^\w\-]', '_', platform_display)
    stamp: str     = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename: str  = f"Scalar_Modifiers_{safe_name}_{stamp}.xlsx"
    filepath: Path = _LOGS_ROOT / filename

    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter

        def _xl(v):
            """Coerce value to openpyxl-safe type."""
            if v is None or isinstance(v, (int, float, str, bool)):
                return v
            try:
                return int(v)
            except (TypeError, ValueError):
                return str(v)

        wb = Workbook()

        # ── Sheet 1: Scalar Modifiers ──────────────────────────────────────
        ws = wb.active
        ws.title = 'Scalar Modifiers'

        COLS: list[str]   = ['Label', 'Type', 'Register', 'Encoding',
                              'Value (Dec)', 'Value (Hex)', 'Converted (MHz/mV)',
                              'Active', 'Fuse Path', 'Description']
        WIDTHS: list[int] = [36, 16, 48, 14, 13, 14, 20, 8, 42, 80]

        hdr_fill = PatternFill(start_color='0071C5', end_color='0071C5', fill_type='solid')
        hdr_font = Font(color='FFFFFF', bold=True, size=11)
        ctr_aln  = Alignment(horizontal='center', vertical='center')
        top_aln  = Alignment(vertical='top')
        wrap_aln = Alignment(wrap_text=True, vertical='top')

        # Type → colour mapping (active registers get a more vivid tint)
        _TYPE_COLOURS: dict[str, tuple[str, str]] = {
            # (active_colour, inactive_colour)
            'p0_override': ('C7D9FF', 'E6F0FF'),
            'itd_voltage':  ('FFD9C0', 'FFF0E6'),
            'itd_slope':    ('FFE8B0', 'FFF8E6'),
            'downbin':      ('C7F0C7', 'F0FFE6'),
            'mct_delta':    ('DFC7FF', 'F0E6FF'),
            'atom_delta':   ('B0F5E8', 'E6FFFA'),
            'acode_min':    ('FFFF99', 'FFFFE6'),
        }
        _DEFAULT_COLOURS = ('E0E0E0', 'F5F5F5')

        for ci, (col_name, width) in enumerate(zip(COLS, WIDTHS), 1):
            cell = ws.cell(row=1, column=ci, value=col_name)
            cell.fill      = hdr_fill
            cell.font      = hdr_font
            cell.alignment = ctr_aln
            ws.column_dimensions[get_column_letter(ci)].width = width
        ws.row_dimensions[1].height = 22
        ws.freeze_panes = 'A2'

        for ri, (reg_name, entry) in enumerate(sorted(scalars.items()), 2):
            # Look up live value + conversion from the discovery cache
            cache_r   = _cache_lookup.get(entry.get('register', reg_name), {})
            raw_val   = cache_r.get('value')
            hex_val   = cache_r.get('hex', '') or (
                f'0x{int(raw_val):X}' if isinstance(raw_val, int) else '')
            converted = cache_r.get('converted', '')
            is_active = bool(cache_r.get('active', False)) if cache_r else False

            stype     = entry.get('type', '')
            act_col, inact_col = _TYPE_COLOURS.get(stype, _DEFAULT_COLOURS)
            hex_col   = act_col if is_active else inact_col
            rfill     = PatternFill(start_color=hex_col, end_color=hex_col, fill_type='solid')

            vals = [
                entry.get('label',       reg_name),
                stype,
                entry.get('register',    reg_name),
                entry.get('encoding',    ''),
                _xl(raw_val),
                hex_val,
                converted,
                'Yes' if is_active else 'No',
                entry.get('fuse_path',   ''),
                entry.get('description', ''),
            ]
            for ci, val in enumerate(vals, 1):
                cell = ws.cell(row=ri, column=ci, value=_xl(val))
                cell.fill      = rfill
                cell.alignment = wrap_aln if ci == len(COLS) else top_aln

        ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"

        # ── Sheet 2: By Type ───────────────────────────────────────────────
        ws2 = wb.create_sheet(title='By Type')
        ws2['A1'] = 'Scalar Modifiers — Grouped by Type'
        ws2['A1'].font = Font(bold=True, size=14, color='0071C5')

        bold = Font(bold=True)
        meta = [
            ('Platform',  platform_display),
            ('Generated', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
            ('Total',     len(scalars)),
            ('Active',    sum(1 for e in scalars.values()
                              if _cache_lookup.get(
                                  e.get('register', ''), {}).get('active'))),
        ]
        for ri, (k, v) in enumerate(meta, 3):
            ws2.cell(row=ri, column=1, value=k).font = bold
            ws2.cell(row=ri, column=2, value=v)

        ws2.cell(row=8, column=1, value='Type').font      = bold
        ws2.cell(row=8, column=2, value='Count').font     = bold
        ws2.cell(row=8, column=3, value='Registers').font = bold

        by_type: dict = {}
        for entry in scalars.values():
            by_type.setdefault(entry.get('type', 'unknown'), []).append(
                entry.get('register', ''))
        for ri, (t, regs) in enumerate(sorted(by_type.items()), 9):
            ws2.cell(row=ri, column=1, value=t)
            ws2.cell(row=ri, column=2, value=len(regs))
            ws2.cell(row=ri, column=3, value=', '.join(regs))

        for col_l, w in [('A', 20), ('B', 8), ('C', 80)]:
            ws2.column_dimensions[col_l].width = w

        wb.save(str(filepath))
        log.info(f"Scalar modifiers exported: {filepath}")
        return str(filepath)

    except Exception as e:
        log.error(f"Scalar modifiers export failed: {e}")
        traceback.print_exc()
        return None


def export_register_change_to_excel(written_list: list, platform_display: str = None) -> str:
    """Export a before/after comparison for one or more discovered register writes.

    Produces the same style of before/after workbook as bump/flatten operations.

    Filename  :  Register_Edit_{Platform}_{YYYYMMDD_HHMMSS}.xlsx
    Location  :  Logs/  (project root)
    Sheets    :
      'Changes'   — one row per register: before / after / delta / verified
      'Summary'   — platform, timestamp, totals

    Args:
        written_list: List of write-result dicts from apply_discovered_register_edits().
                      Each dict must contain keys: reg_name, fuse_path,
                      before (int), after (int), verified (bool).
        platform_display: Human-readable platform name.

    Returns:
        Absolute path to the created file, or None on failure.
    """
    if not written_list:
        return None

    if platform_display is None:
        _, _, platform_display, _ = load_discovery_cache()
    if platform_display is None:
        platform_display = 'Unknown_Platform'

    # Enrich each row with domain/category from cache (best-effort)
    cached_recs, _, _, _ = load_discovery_cache()
    _lookup: dict = {}
    if cached_recs:
        for r in cached_recs:
            _lookup[r['name']] = r

    safe_name: str = re.sub(r'[^\w\-]', '_', platform_display)
    stamp: str     = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename: str  = f"Register_Edit_{safe_name}_{stamp}.xlsx"
    filepath: Path = _LOGS_ROOT / filename

    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter

        wb = Workbook()

        # ── Sheet 1: Changes ───────────────────────────────────────────────
        ws = wb.active
        ws.title = 'Changes'

        COLS: list[str]   = ['Register Name', 'Domain', 'Category', 'Fuse Path',
                              'Before (Dec)', 'Before (Hex)',
                              'After (Dec)',  'After (Hex)',
                              'Δ (After−Before)', 'Verified']
        WIDTHS: list[int] = [52, 14, 16, 38, 13, 14, 13, 14, 18, 10]

        hdr_fill = PatternFill(start_color='0071C5', end_color='0071C5', fill_type='solid')
        hdr_font = Font(color='FFFFFF', bold=True, size=11)
        ok_fill  = PatternFill(start_color='E6F4EA', end_color='E6F4EA', fill_type='solid')
        mis_fill = PatternFill(start_color='FDECEA', end_color='FDECEA', fill_type='solid')
        ctr_aln  = Alignment(horizontal='center', vertical='center')
        top_aln  = Alignment(vertical='center')

        for ci, (col_name, width) in enumerate(zip(COLS, WIDTHS), 1):
            cell = ws.cell(row=1, column=ci, value=col_name)
            cell.fill      = hdr_fill
            cell.font      = hdr_font
            cell.alignment = ctr_aln
            ws.column_dimensions[get_column_letter(ci)].width = width
        ws.row_dimensions[1].height = 22
        ws.freeze_panes = 'A2'

        for ri, w in enumerate(written_list, 2):
            verified = bool(w.get('verified', False))
            rfill    = ok_fill if verified else mis_fill
            cache_r  = _lookup.get(w.get('reg_name', ''), {})
            bval     = w.get('before')
            aval     = w.get('after')
            delta    = (aval - bval) if (isinstance(bval, int) and isinstance(aval, int)) else ''
            vals = [
                w.get('reg_name', ''),
                cache_r.get('domain',   ''),
                cache_r.get('category', ''),
                w.get('fuse_path', '') or cache_r.get('fuse_path', ''),
                bval,
                f'0x{bval:X}' if isinstance(bval, int) else '',
                aval,
                f'0x{aval:X}' if isinstance(aval, int) else '',
                delta,
                'Yes ✓' if verified else 'MISMATCH ✗',
            ]
            for ci, val in enumerate(vals, 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.fill      = rfill
                cell.alignment = top_aln

        ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"

        # ── Sheet 2: Summary ───────────────────────────────────────────────
        ws2 = wb.create_sheet(title='Summary')
        ws2['A1'] = 'Register Edit Summary'
        ws2['A1'].font = Font(bold=True, size=14, color='0071C5')

        bold = Font(bold=True)
        verified_count = sum(1 for w in written_list if w.get('verified'))
        meta = [
            ('Platform',         platform_display),
            ('Timestamp',        datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
            ('Registers Written', len(written_list)),
            ('Verified OK',      verified_count),
            ('Mismatches',       len(written_list) - verified_count),
        ]
        for ri, (k, v) in enumerate(meta, 3):
            ws2.cell(row=ri, column=1, value=k).font = bold
            ws2.cell(row=ri, column=2, value=v)

        for col_l, w in [('A', 22), ('B', 30)]:
            ws2.column_dimensions[col_l].width = w

        wb.save(str(filepath))
        log.info(f"Register edit exported: {filepath}")
        return str(filepath)

    except Exception as e:
        log.error(f"Register change export failed: {e}")
        traceback.print_exc()
        return None


def export_scalar_change_to_excel(reg_name: str, before: dict, after: dict,
                                   info: dict, platform_display: str = None) -> str:
    """Export a before/after comparison for a single scalar modifier write.

    Filename  :  Scalar_Edit_{Platform}_{YYYYMMDD_HHMMSS}.xlsx
    Location  :  Logs/  (project root)
    Sheets    :
      'Change'  — one row for the edited register with before / after columns
      'Summary' — platform, timestamp, register metadata

    Args:
        reg_name:         Register key (name in vf_domains.json scalar_modifiers).
        before:           Return dict from read_scalar_modifier() before the write.
        after:            Return dict from read_scalar_modifier() after the write.
        info:             Entry dict from scalar_modifiers config (type, label, encoding…).
        platform_display: Human-readable platform name.

    Returns:
        Absolute path to the created file, or None on failure.
    """
    if platform_display is None:
        _, _, platform_display, _ = load_discovery_cache()
    if platform_display is None:
        platform_display = 'Unknown_Platform'

    safe_name: str = re.sub(r'[^\w\-]', '_', platform_display)
    stamp: str     = datetime.now().strftime('%Y%m%d_%H%M%S')
    safe_reg: str  = re.sub(r'[^\w\-]', '_', reg_name)[:40]
    filename: str  = f"Scalar_Edit_{safe_name}_{safe_reg}_{stamp}.xlsx"
    filepath: Path = _LOGS_ROOT / filename

    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        label   = info.get('label', reg_name)
        units   = before.get('units', after.get('units', 'raw'))
        braw    = before.get('raw')
        araw    = after.get('raw')
        bconv   = before.get('converted')
        aconv   = after.get('converted')
        delta   = (aconv - bconv) if (isinstance(bconv, (int, float)) and
                                       isinstance(aconv, (int, float))) else ''

        # ── Sheet 1: Change ────────────────────────────────────────────────
        ws = wb.active
        ws.title = 'Change'

        COLS: list[str]   = ['Label', 'Type', 'Register', 'Encoding',
                              'Before (Raw)', f'Before ({units})',
                              'After (Raw)',  f'After ({units})',
                              f'Δ ({units})', 'Write OK']
        WIDTHS: list[int] = [36, 16, 48, 14, 13, 18, 13, 18, 14, 10]

        hdr_fill = PatternFill(start_color='0071C5', end_color='0071C5', fill_type='solid')
        hdr_font = Font(color='FFFFFF', bold=True, size=11)
        ok_fill  = PatternFill(start_color='E6F4EA', end_color='E6F4EA', fill_type='solid')
        ctr_aln  = Alignment(horizontal='center', vertical='center')
        top_aln  = Alignment(vertical='center')

        for ci, (col_name, width) in enumerate(zip(COLS, WIDTHS), 1):
            cell = ws.cell(row=1, column=ci, value=col_name)
            cell.fill      = hdr_fill
            cell.font      = hdr_font
            cell.alignment = ctr_aln
            ws.column_dimensions[get_column_letter(ci)].width = width
        ws.row_dimensions[1].height = 22
        ws.freeze_panes = 'A2'

        row_vals = [
            label,
            info.get('type', ''),
            info.get('register', reg_name),
            info.get('encoding', ''),
            braw,  bconv,
            araw,  aconv,
            delta,
            'Yes ✓',
        ]
        for ci, val in enumerate(row_vals, 1):
            cell = ws.cell(row=2, column=ci, value=val)
            cell.fill      = ok_fill
            cell.alignment = top_aln

        # ── Sheet 2: Summary ───────────────────────────────────────────────
        ws2 = wb.create_sheet(title='Summary')
        ws2['A1'] = 'Scalar Modifier Edit Summary'
        ws2['A1'].font = Font(bold=True, size=14, color='0071C5')

        bold = Font(bold=True)
        meta = [
            ('Platform',   platform_display),
            ('Timestamp',  datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
            ('Label',      label),
            ('Type',       info.get('type', '')),
            ('Register',   info.get('register', reg_name)),
            ('Encoding',   info.get('encoding', '')),
            ('Fuse Path',  info.get('fuse_path', '')),
            ('Units',      units),
            ('Before',     f'{bconv} {units}  (raw={braw})'),
            ('After',      f'{aconv} {units}  (raw={araw})'),
            ('Delta',      f'{delta} {units}' if delta != '' else 'n/a'),
        ]
        for ri, (k, v) in enumerate(meta, 3):
            ws2.cell(row=ri, column=1, value=k).font = bold
            ws2.cell(row=ri, column=2, value=str(v))

        for col_l, w in [('A', 16), ('B', 50)]:
            ws2.column_dimensions[col_l].width = w

        wb.save(str(filepath))
        log.info(f"Scalar edit exported: {filepath}")
        return str(filepath)

    except Exception as e:
        log.error(f"Scalar change export failed: {e}")
        traceback.print_exc()
        return None


def _count_active(all_path_results: dict) -> int:
    """Count total active (non-zero) registers across all fuse paths."""
    return sum(
        len([r for r in regs if r['active']])
        for results_by_cat in all_path_results.values()
        for regs in results_by_cat.values()
    )

